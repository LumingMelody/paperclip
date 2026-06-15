#!/usr/bin/env python3
"""Build and optionally push the EP Shopify weekly return-rate report.

The script mirrors ``weekly_return_report.py`` operationally: direct DWS reads,
markdown rendering, DingTalk ``groupMessages/send`` push, and argparse CLI.
By default it is a dry run and prints markdown. Pass ``--send`` to push.
"""
from __future__ import annotations

import argparse
import json
import os
import plistlib
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

import httpx
import pymysql
from loguru import logger
from pymysql.cursors import DictCursor


DEFAULT_SITES = ("UK", "US", "FR", "DE")
DEFAULT_CHANNEL = "concierge"
DEFAULT_MATURITY_DAYS = 45
DEFAULT_TOP_STYLES = 10
DEFAULT_MIN_TOP_STYLE_QTY = 50
DEFAULT_LOW_SAMPLE_SALES_QTY = 1000
STYLE_TYPE_ORDER = ("迭代前", "迭代后", "新款", "pre-order", "老款(未分类)")
FULL_STYLE_TYPE_ORDER = ("老款", "迭代前", "迭代后", "新款", "pre-order")
STYLE_TAG_PATH = Path(__file__).parent / "site_style_tag.json"
WAREHOUSE_MAP_PATH = Path(__file__).parent / "site_warehouse_map.json"
OUTPUT_DIR = Path(__file__).parent / "output"
WAREHOUSE_DISPLAY_ORDER = ("CN发货", "万州仓发货", "US仓发货", "谷仓发货", "天马发货", "科隆仓发货", "未映射", "无仓库记录")
DEFAULT_FULL_TIMING_MATURITY_DAYS = 120
DEFAULT_PREDICTION_MIN_RETURN_QTY = 10
PREDICTION_PROGRESS_FLOOR = 0.30
PREDICTION_RATE_CAP = 0.95

TOKEN_URL = "https://api.dingtalk.com/v1.0/oauth2/accessToken"
SEND_URL = "https://api.dingtalk.com/v1.0/robot/groupMessages/send"

_REQUIRED_DWS_ENV = ("DWS_DB_HOST", "DWS_DB_USER", "DWS_DB_PASSWORD", "DWS_DB_DATABASE")
_SITE_RE = re.compile(r"^(US|UK|FR|DE)$")


@dataclass(frozen=True)
class WeekWindow:
    since: str
    until: str | None
    maturity_days: int


@dataclass(frozen=True)
class StyleTag:
    style_type: str
    primary_category: str | None


@dataclass(frozen=True)
class DingTalkTarget:
    app_key: str
    app_secret: str
    open_conversation_id: str
    robot_code: str | None


@dataclass(frozen=True)
class SiteReportData:
    site: str
    account: str
    metadata: dict[str, Any]
    summary: dict[str, Any]
    style_type_rows: list[dict[str, Any]]
    top_styles: list[dict[str, Any]]
    timing_rows: list[dict[str, Any]]
    order_unit_rows: list[dict[str, Any]]
    warehouse_rows: list[dict[str, Any]]
    dirty_warehouse_pct: float


@dataclass(frozen=True)
class ReportData:
    since: str
    until: str | None
    maturity_days: int
    sites: list[SiteReportData]


@dataclass(frozen=True)
class PredictionResult:
    rate: float | None
    progress: float
    curve_level: str
    low_confidence: bool


@dataclass(frozen=True)
class FullSiteReportData:
    site: str
    account: str
    metadata: dict[str, Any]
    summary: dict[str, Any]
    cohort_age_days: int | None
    table1_rows: list[dict[str, Any]]
    table2_rows: list[dict[str, Any]]
    order_unit_rows: list[dict[str, Any]]
    warehouse_rows: list[dict[str, Any]]
    dirty_warehouse_pct: float
    timing_metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(frozen=True)
class FullReportData:
    since: str
    until: str | None
    maturity_days: int
    timing_maturity_days: int
    sites: list[FullSiteReportData]


def site_to_account(site: str) -> str:
    site = site.upper()
    if not _SITE_RE.match(site):
        raise ValueError(f"site must be one of {', '.join(DEFAULT_SITES)}, got {site!r}")
    return f"EPSITE{site}"


def _parse_iso_date(value: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise argparse.ArgumentTypeError(f"expected YYYY-MM-DD, got {value!r}") from exc


def default_window(today: date | None = None, maturity_days: int = DEFAULT_MATURITY_DAYS) -> WeekWindow:
    today = today or date.today()
    until = today - timedelta(days=maturity_days)
    since = until - timedelta(days=7)
    return WeekWindow(since=since.isoformat(), until=None, maturity_days=maturity_days)


def resolve_window(since_arg: str | None, until_arg: str | None, maturity_days: int) -> WeekWindow:
    default = default_window(maturity_days=maturity_days)
    since = _parse_iso_date(since_arg).isoformat() if since_arg else default.since
    until = _parse_iso_date(until_arg).isoformat() if until_arg else None
    if until and since >= until:
        raise ValueError(f"--since must be before --until, got {since} >= {until}")
    return WeekWindow(since=since, until=until, maturity_days=maturity_days)


def _serialize(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _serialize_row(row: dict[str, Any]) -> dict[str, Any]:
    return {key: _serialize(value) for key, value in row.items()}


def _int(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, Decimal):
        return int(value)
    return int(value)


def _float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    return float(value)


def _rate(numerator: float, denominator: float) -> float | None:
    if denominator <= 0:
        return None
    return round(numerator / denominator, 4)


def _pct(value: float | None, digits: int = 1) -> str:
    if value is None:
        return "-"
    return f"{value * 100:.{digits}f}%"


def _load_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError:
        return {}


def _find_normalized(obj: Any, names: set[str]) -> str | None:
    if isinstance(obj, dict):
        for key, value in obj.items():
            normalized = str(key).replace("-", "_").lower()
            if normalized in names and value:
                return str(value)
        for value in obj.values():
            found = _find_normalized(value, names)
            if found:
                return found
    elif isinstance(obj, list):
        for item in obj:
            found = _find_normalized(item, names)
            if found:
                return found
    return None


def _find_in_named_section(obj: Any, section_hint: str, names: set[str]) -> str | None:
    if isinstance(obj, dict):
        for key, value in obj.items():
            normalized = str(key).replace("-", "_").lower()
            if section_hint in normalized:
                found = _find_normalized(value, names)
                if found:
                    return found
            found = _find_in_named_section(value, section_hint, names)
            if found:
                return found
    elif isinstance(obj, list):
        for item in obj:
            found = _find_in_named_section(item, section_hint, names)
            if found:
                return found
    return None


def _read_dotenv(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    try:
        lines = path.read_text(encoding="utf-8").splitlines()
    except FileNotFoundError:
        return out
    for line in lines:
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        out[key.strip()] = value.strip().strip('"').strip("'")
    return out


def _load_tool_secrets() -> dict[str, Any]:
    path = Path(os.path.expanduser("~/.paperclip/tool-secrets.json"))
    payload = _load_json(path)
    return payload if isinstance(payload, dict) else {}


def _resolve_dws_config() -> dict[str, str]:
    env_config = {
        "host": os.environ.get("DWS_DB_HOST"),
        "port": os.environ.get("DWS_DB_PORT"),
        "user": os.environ.get("DWS_DB_USER"),
        "password": os.environ.get("DWS_DB_PASSWORD"),
        "database": os.environ.get("DWS_DB_DATABASE"),
    }
    if all(env_config[key] for key in ("host", "user", "password", "database")):
        return {
            "host": str(env_config["host"]),
            "port": str(env_config["port"] or "3306"),
            "user": str(env_config["user"]),
            "password": str(env_config["password"]),
            "database": str(env_config["database"]),
        }

    secrets = _load_tool_secrets()
    dws: dict[str, Any] | None = None
    try:
        companies = secrets.get("companies") or {}
        if isinstance(companies, dict) and companies:
            first_company = next(iter(companies.values()))
            candidate = first_company.get("dws") if isinstance(first_company, dict) else None
            if isinstance(candidate, dict):
                dws = candidate
    except Exception:
        dws = None

    if dws:
        return {
            "host": str(env_config["host"] or dws.get("host") or ""),
            "port": str(env_config["port"] or dws.get("port") or "3306"),
            "user": str(env_config["user"] or dws.get("user") or ""),
            "password": str(env_config["password"] or dws.get("password") or ""),
            "database": str(env_config["database"] or dws.get("database") or ""),
        }

    missing = [name for name in _REQUIRED_DWS_ENV if not os.environ.get(name)]
    raise RuntimeError(
        "missing DWS credentials: "
        + ", ".join(missing)
        + " (or configure ~/.paperclip/tool-secrets.json companies.*.dws)"
    )


def _connect() -> pymysql.Connection:
    cfg = _resolve_dws_config()
    missing = [key for key in ("host", "user", "password", "database") if not cfg.get(key)]
    if missing:
        raise RuntimeError(f"missing DWS config fields: {', '.join(missing)}")
    return pymysql.connect(
        host=cfg["host"],
        port=int(cfg.get("port") or "3306"),
        user=cfg["user"],
        password=cfg["password"],
        database=cfg["database"],
        charset="utf8mb4",
        connect_timeout=8,
        cursorclass=DictCursor,
    )


def _fetch_one(conn: pymysql.Connection, sql: str, params: dict[str, Any]) -> dict[str, Any]:
    with conn.cursor() as cur:
        cur.execute(sql, params)
        row = cur.fetchone()
    return _serialize_row(dict(row or {}))


def _fetch_all(conn: pymysql.Connection, sql: str, params: dict[str, Any]) -> list[dict[str, Any]]:
    with conn.cursor() as cur:
        cur.execute(sql, params)
        return [_serialize_row(dict(row)) for row in cur.fetchall()]


def cohort_metadata(
    conn: pymysql.Connection,
    since: str,
    until: str | None,
    maturity_days: int,
) -> tuple[str, dict[str, Any]]:
    effective_until = "%(until)s" if until else "DATE_SUB(CURDATE(), INTERVAL %(maturity_days)s DAY)"
    params: dict[str, Any] = {"since": since, "maturity_days": maturity_days}
    if until:
        params["until"] = until
    metadata_sql = f"""
        SELECT
            CURDATE() AS asOfDate,
            %(since)s AS windowStart,
            CAST({effective_until} AS DATE) AS windowEnd,
            CAST(DATE_SUB(CAST({effective_until} AS DATE), INTERVAL 1 DAY) AS DATE) AS coveredThrough,
            %(maturity_days)s AS maturityDays,
            CASE
                WHEN {effective_until} > DATE_SUB(CURDATE(), INTERVAL %(maturity_days)s DAY) THEN TRUE
                ELSE FALSE
            END AS windowIncludesImmature
    """
    metadata = _fetch_one(conn, metadata_sql, params)
    metadata["maturityDays"] = _int(metadata.get("maturityDays"))
    metadata["windowIncludesImmature"] = bool(metadata.get("windowIncludesImmature"))
    return effective_until, metadata


def fetch_overall_summary(
    conn: pymysql.Connection,
    account: str,
    since: str,
    until: str | None,
    maturity_days: int,
) -> tuple[dict[str, Any], dict[str, Any]]:
    effective_until, metadata = cohort_metadata(conn, since, until, maturity_days)
    sql = f"""
        SELECT
            COUNT(*) AS rowCount,
            COUNT(DISTINCT orderid) AS orderCount,
            CAST(COALESCE(SUM(quantity), 0) AS DECIMAL(20,0)) AS salesQty,
            CAST(COALESCE(SUM(COALESCE(return_quantity, 0)), 0) AS DECIMAL(20,0)) AS returnQty
        FROM dm_od_shopify_resreturn_d
        WHERE account = %(account)s
          AND pay_time >= %(since)s
          AND pay_time < {effective_until}
    """
    params: dict[str, Any] = {
        "account": account,
        "since": since,
        "maturity_days": maturity_days,
    }
    if until:
        params["until"] = until
    row = _fetch_one(conn, sql, params)
    sales_qty = _float(row.get("salesQty"))
    return_qty = _float(row.get("returnQty"))
    return {
        "rowCount": _int(row.get("rowCount")),
        "orderCount": _int(row.get("orderCount")),
        "salesQty": sales_qty,
        "returnQty": return_qty,
        "returnRate": _rate(return_qty, sales_qty),
    }, metadata


def fetch_site_return_rate_by_style(
    conn: pymysql.Connection,
    account: str,
    since: str,
    until: str | None,
    maturity_days: int,
    top: int | None = None,
    min_qty: int | None = None,
) -> list[dict[str, Any]]:
    effective_until, _metadata = cohort_metadata(conn, since, until, maturity_days)
    sql = f"""
        SELECT
            LEFT(shipping_sku, 7) AS styleCode,
            CAST(COALESCE(SUM(quantity), 0) AS DECIMAL(20,0)) AS salesQty,
            CAST(COALESCE(SUM(COALESCE(return_quantity, 0)), 0) AS DECIMAL(20,0)) AS returnQty,
            COUNT(DISTINCT shipping_sku) AS skuCount
        FROM dm_od_shopify_resreturn_d
        WHERE account = %(account)s
          AND pay_time >= %(since)s
          AND pay_time < {effective_until}
          AND shipping_sku IS NOT NULL
          AND shipping_sku != ''
    """
    params: dict[str, Any] = {
        "account": account,
        "since": since,
        "maturity_days": maturity_days,
    }
    if until:
        params["until"] = until
    sql += " GROUP BY LEFT(shipping_sku, 7)"
    if top is not None:
        params["top"] = top
        params["min_qty"] = min_qty if min_qty is not None else DEFAULT_MIN_TOP_STYLE_QTY
        sql += """
            HAVING salesQty >= %(min_qty)s
            ORDER BY (SUM(COALESCE(return_quantity, 0)) / NULLIF(SUM(quantity), 0)) DESC
            LIMIT %(top)s
        """
    else:
        sql += " ORDER BY LEFT(shipping_sku, 7)"
    rows = _fetch_all(conn, sql, params)
    out: list[dict[str, Any]] = []
    for row in rows:
        sales_qty = _float(row.get("salesQty"))
        return_qty = _float(row.get("returnQty"))
        row["salesQty"] = sales_qty
        row["returnQty"] = return_qty
        row["skuCount"] = _int(row.get("skuCount"))
        row["returnRate"] = _rate(return_qty, sales_qty)
        out.append(row)
    return out


def fetch_site_return_timing_by_style(
    conn: pymysql.Connection,
    account: str,
    since: str,
    until: str | None,
    maturity_days: int,
) -> list[dict[str, Any]]:
    rows, _metadata = fetch_site_return_timing_by_style_with_metadata(
        conn,
        account,
        since,
        until,
        maturity_days,
    )
    return rows


def fetch_site_return_timing_by_style_with_metadata(
    conn: pymysql.Connection,
    account: str,
    since: str,
    until: str | None,
    maturity_days: int,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    effective_until, metadata = cohort_metadata(conn, since, until, maturity_days)
    sql = f"""
        SELECT
            LEFT(shipping_sku, 7) AS styleCode,
            CAST(COALESCE(SUM(COALESCE(return_quantity, 0)), 0) AS DECIMAL(20,0)) AS returnedQty,
            CAST(COALESCE(SUM(CASE
                WHEN DATEDIFF(return_time, pay_time) <= 30 THEN COALESCE(return_quantity, 0)
                ELSE 0
            END), 0) AS DECIMAL(20,0)) AS qty_0_30,
            CAST(COALESCE(SUM(CASE
                WHEN DATEDIFF(return_time, pay_time) BETWEEN 31 AND 45 THEN COALESCE(return_quantity, 0)
                ELSE 0
            END), 0) AS DECIMAL(20,0)) AS qty_31_45,
            CAST(COALESCE(SUM(CASE
                WHEN DATEDIFF(return_time, pay_time) > 45 THEN COALESCE(return_quantity, 0)
                ELSE 0
            END), 0) AS DECIMAL(20,0)) AS qty_45plus
        FROM dm_od_shopify_resreturn_d
        WHERE account = %(account)s
          AND pay_time >= %(since)s
          AND pay_time < {effective_until}
          AND return_quantity > 0
          AND return_time IS NOT NULL
          AND pay_time IS NOT NULL
          AND shipping_sku IS NOT NULL
          AND shipping_sku != ''
        GROUP BY LEFT(shipping_sku, 7)
        ORDER BY returnedQty DESC
    """
    params: dict[str, Any] = {
        "account": account,
        "since": since,
        "maturity_days": maturity_days,
    }
    if until:
        params["until"] = until
    rows = _fetch_all(conn, sql, params)
    out: list[dict[str, Any]] = []
    for row in rows:
        returned_qty = _float(row.get("returnedQty"))
        row["returnedQty"] = returned_qty
        for key in ("qty_0_30", "qty_31_45", "qty_45plus"):
            row[key] = _float(row.get(key))
        row["pct_0_30"] = _rate(row["qty_0_30"], returned_qty)
        row["pct_31_45"] = _rate(row["qty_31_45"], returned_qty)
        row["pct_45plus"] = _rate(row["qty_45plus"], returned_qty)
        out.append(row)
    return out, metadata


def fetch_site_return_rate_by_order_units(
    conn: pymysql.Connection,
    account: str,
    since: str,
    until: str | None,
    maturity_days: int,
) -> list[dict[str, Any]]:
    effective_until, _metadata = cohort_metadata(conn, since, until, maturity_days)
    sql = f"""
        WITH order_totals AS (
            SELECT
                orderid,
                CAST(COALESCE(SUM(quantity), 0) AS DECIMAL(20,0)) AS orderSalesQty,
                CAST(COALESCE(SUM(COALESCE(return_quantity, 0)), 0) AS DECIMAL(20,0)) AS orderReturnQty
            FROM dm_od_shopify_resreturn_d
            WHERE account = %(account)s
              AND pay_time >= %(since)s
              AND pay_time < {effective_until}
              AND orderid IS NOT NULL
              AND orderid != ''
            GROUP BY orderid
        ),
        bucketed AS (
            SELECT
                CASE
                    WHEN orderSalesQty >= 5 THEN '5+'
                    ELSE CAST(orderSalesQty AS CHAR)
                END AS unitsBucket,
                orderSalesQty,
                orderReturnQty
            FROM order_totals
            WHERE orderSalesQty BETWEEN 1 AND 4 OR orderSalesQty >= 5
        )
        SELECT
            unitsBucket,
            COUNT(*) AS orderCount,
            CAST(COALESCE(SUM(orderSalesQty), 0) AS DECIMAL(20,0)) AS salesQty,
            CAST(COALESCE(SUM(orderReturnQty), 0) AS DECIMAL(20,0)) AS returnQty
        FROM bucketed
        GROUP BY unitsBucket
        ORDER BY FIELD(unitsBucket, '1', '2', '3', '4', '5+')
    """
    params: dict[str, Any] = {"account": account, "since": since, "maturity_days": maturity_days}
    if until:
        params["until"] = until
    rows = _fetch_all(conn, sql, params)
    out: list[dict[str, Any]] = []
    for row in rows:
        sales_qty = _float(row.get("salesQty"))
        return_qty = _float(row.get("returnQty"))
        row["orderCount"] = _int(row.get("orderCount"))
        row["salesQty"] = sales_qty
        row["returnQty"] = return_qty
        row["returnRate"] = _rate(return_qty, sales_qty)
        out.append(row)
    return out


def fetch_site_return_rate_by_warehouse(
    conn: pymysql.Connection,
    account: str,
    since: str,
    until: str | None,
    maturity_days: int,
) -> tuple[list[dict[str, Any]], float]:
    effective_until, _metadata = cohort_metadata(conn, since, until, maturity_days)
    sql = f"""
        WITH src AS (
            SELECT
                CASE
                    WHEN warehouseName IS NULL OR TRIM(warehouseName) = '' OR warehouseName = '无仓库记录'
                    THEN '无仓库记录'
                    ELSE warehouseName
                END AS warehouseName,
                quantity,
                COALESCE(return_quantity, 0) AS return_quantity,
                CASE
                    WHEN warehouseName IS NULL OR TRIM(warehouseName) = '' OR warehouseName = '无仓库记录'
                    THEN 1
                    ELSE 0
                END AS isDirtyWarehouse
            FROM dm_od_shopify_resreturn_d
            WHERE account = %(account)s
              AND pay_time >= %(since)s
              AND pay_time < {effective_until}
        ),
        totals AS (
            SELECT
                CAST(COALESCE(SUM(return_quantity), 0) AS DECIMAL(20,0)) AS totalReturnQty,
                COUNT(*) AS totalRows,
                COALESCE(SUM(isDirtyWarehouse), 0) AS dirtyRows
            FROM src
        ),
        grouped AS (
            SELECT
                warehouseName,
                CAST(COALESCE(SUM(quantity), 0) AS DECIMAL(20,0)) AS salesQty,
                CAST(COALESCE(SUM(return_quantity), 0) AS DECIMAL(20,0)) AS returnQty
            FROM src
            GROUP BY warehouseName
        )
        SELECT
            g.warehouseName,
            g.salesQty,
            g.returnQty,
            CASE
                WHEN t.totalReturnQty > 0 THEN g.returnQty / t.totalReturnQty
                ELSE 0
            END AS returnShare,
            CASE
                WHEN t.totalRows > 0 THEN t.dirtyRows / t.totalRows
                ELSE 0
            END AS dirtyWarehousePct
        FROM grouped g
        CROSS JOIN totals t
        ORDER BY g.returnQty DESC, g.salesQty DESC
    """
    params: dict[str, Any] = {"account": account, "since": since, "maturity_days": maturity_days}
    if until:
        params["until"] = until
    rows = _fetch_all(conn, sql, params)
    out: list[dict[str, Any]] = []
    dirty_warehouse_pct = 0.0
    for row in rows:
        sales_qty = _float(row.get("salesQty"))
        return_qty = _float(row.get("returnQty"))
        dirty_warehouse_pct = round(_float(row.pop("dirtyWarehousePct")), 4)
        row["salesQty"] = sales_qty
        row["returnQty"] = return_qty
        row["returnRate"] = _rate(return_qty, sales_qty)
        row["returnShare"] = round(_float(row.get("returnShare")), 4)
        out.append(row)
    return out, dirty_warehouse_pct


def fetch_site_return_rate_by_raw_warehouse(
    conn: pymysql.Connection,
    account: str,
    since: str,
    until: str | None,
    maturity_days: int,
) -> list[dict[str, Any]]:
    effective_until, _metadata = cohort_metadata(conn, since, until, maturity_days)
    sql = f"""
        SELECT
            warehouseName AS rawWarehouseName,
            CAST(COALESCE(SUM(quantity), 0) AS DECIMAL(20,0)) AS salesQty,
            CAST(COALESCE(SUM(COALESCE(return_quantity, 0)), 0) AS DECIMAL(20,0)) AS returnQty
        FROM dm_od_shopify_resreturn_d
        WHERE account = %(account)s
          AND pay_time >= %(since)s
          AND pay_time < {effective_until}
        GROUP BY warehouseName
    """
    params: dict[str, Any] = {"account": account, "since": since, "maturity_days": maturity_days}
    if until:
        params["until"] = until
    rows = _fetch_all(conn, sql, params)
    out: list[dict[str, Any]] = []
    for row in rows:
        sales_qty = _float(row.get("salesQty"))
        return_qty = _float(row.get("returnQty"))
        row["salesQty"] = sales_qty
        row["returnQty"] = return_qty
        row["returnRate"] = _rate(return_qty, sales_qty)
        out.append(row)
    return out


def load_style_tags(path: Path = STYLE_TAG_PATH) -> dict[str, StyleTag]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    styles = payload.get("styles") if isinstance(payload, dict) else None
    if not isinstance(styles, dict):
        raise ValueError(f"invalid style tag file: missing styles object at {path}")
    out: dict[str, StyleTag] = {}
    for style, value in styles.items():
        if not isinstance(value, dict):
            continue
        style_type = str(value.get("styleType") or "老款(未分类)")
        if style_type not in STYLE_TYPE_ORDER:
            style_type = "老款(未分类)"
        primary_category = value.get("primaryCategory")
        out[str(style)] = StyleTag(style_type=style_type, primary_category=str(primary_category) if primary_category else None)
    return out


def style_type_for(style_code: str | None, style_tags: dict[str, StyleTag]) -> str:
    if not style_code:
        return "老款(未分类)"
    tag = style_tags.get(style_code)
    return tag.style_type if tag else "老款(未分类)"


def full_style_type_for(style_code: str | None, style_tags: dict[str, StyleTag]) -> str:
    style_type = style_type_for(style_code, style_tags)
    return "老款" if style_type == "老款(未分类)" else style_type


def load_warehouse_map(path: Path = WAREHOUSE_MAP_PATH) -> dict[str, str]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"invalid warehouse map file: expected object at {path}")
    out: dict[str, str] = {}
    for raw, label in payload.items():
        raw_name = str(raw).strip()
        mapped_name = str(label).strip()
        if raw_name and mapped_name:
            out[raw_name] = mapped_name
    return out


def map_warehouse_name(raw_name: Any, warehouse_map: dict[str, str]) -> str:
    if raw_name is None:
        return "无仓库记录"
    name = str(raw_name).strip()
    if not name or name == "无仓库记录":
        return "无仓库记录"
    return warehouse_map.get(name, "未映射")


def _timing_curve_from_totals(returned_qty: float, qty_0_30: float, qty_31_45: float, qty_45plus: float) -> dict[str, Any]:
    return {
        "returnedQty": returned_qty,
        "pct_0_30": _rate(qty_0_30, returned_qty),
        "pct_31_45": _rate(qty_31_45, returned_qty),
        "pct_45plus": _rate(qty_45plus, returned_qty),
    }


def _sum_timing_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    returned_qty = sum(_float(row.get("returnedQty")) for row in rows)
    qty_0_30 = sum(_float(row.get("qty_0_30")) for row in rows)
    qty_31_45 = sum(_float(row.get("qty_31_45")) for row in rows)
    qty_45plus = sum(_float(row.get("qty_45plus")) for row in rows)
    return _timing_curve_from_totals(returned_qty, qty_0_30, qty_31_45, qty_45plus)


def maturity_progress_beta(
    curve: dict[str, Any] | None,
    age_days: int | float | None,
    maturity_days: int = DEFAULT_FULL_TIMING_MATURITY_DAYS,
    progress_floor: float = PREDICTION_PROGRESS_FLOOR,
) -> float:
    if not curve or age_days is None:
        return 1.0
    age = max(0.0, float(age_days))
    maturity = max(46.0, float(maturity_days))
    pct_0_30 = max(0.0, _float(curve.get("pct_0_30")))
    pct_31_45 = max(0.0, _float(curve.get("pct_31_45")))
    progress_at_30 = min(1.0, pct_0_30)
    progress_at_45 = min(1.0, pct_0_30 + pct_31_45)

    if age >= maturity:
        progress = 1.0
    elif age <= 30:
        progress = progress_at_30 * (age / 30.0)
    elif age <= 45:
        progress = progress_at_30 + (progress_at_45 - progress_at_30) * ((age - 30.0) / 15.0)
    else:
        progress = progress_at_45 + (1.0 - progress_at_45) * ((age - 45.0) / (maturity - 45.0))
    return min(1.0, max(progress_floor, progress))


def predict_restored_return_rate_beta(
    current_return_rate: float | None,
    age_days: int | float | None,
    style_curve: dict[str, Any] | None,
    style_type_curve: dict[str, Any] | None,
    site_curve: dict[str, Any] | None,
    *,
    min_returned_qty: int = DEFAULT_PREDICTION_MIN_RETURN_QTY,
    maturity_days: int = DEFAULT_FULL_TIMING_MATURITY_DAYS,
    progress_floor: float = PREDICTION_PROGRESS_FLOOR,
    rate_cap: float = PREDICTION_RATE_CAP,
) -> PredictionResult:
    """Predict restored return rate from observed rate and mature return timing.

    Assumptions: the current cohort's observed return rate is incomplete when its
    effective age is below the selected mature horizon. A mature timing curve
    estimates what share of eventual returns should already be visible at that
    age. Style-level curves are preferred; when a style has fewer mature returned
    units than ``min_returned_qty``, the curve falls back to style type, then
    site. The progress floor limits over-amplification for very young cohorts,
    and ``rate_cap`` prevents impossible/extreme predictions.
    """
    if current_return_rate is None:
        return PredictionResult(rate=None, progress=1.0, curve_level="none", low_confidence=True)

    candidates = (
        ("style", style_curve),
        ("styleType", style_type_curve),
        ("site", site_curve),
    )
    selected_level = "none"
    selected_curve: dict[str, Any] | None = None
    low_confidence = True
    for level, curve in candidates:
        if curve and _float(curve.get("returnedQty")) >= min_returned_qty:
            selected_level = level
            selected_curve = curve
            low_confidence = level != "style"
            break
    if selected_curve is None:
        selected_curve = site_curve or style_type_curve or style_curve
        selected_level = "site" if site_curve else ("styleType" if style_type_curve else ("style" if style_curve else "none"))
        low_confidence = True

    progress = maturity_progress_beta(
        selected_curve,
        age_days,
        maturity_days=maturity_days,
        progress_floor=progress_floor,
    )
    predicted = min(rate_cap, current_return_rate / progress) if progress > 0 else current_return_rate
    return PredictionResult(
        rate=round(predicted, 4),
        progress=round(progress, 4),
        curve_level=selected_level,
        low_confidence=low_confidence,
    )


def rollup_mapped_warehouse_rows(
    raw_rows: list[dict[str, Any]],
    warehouse_map: dict[str, str],
) -> tuple[list[dict[str, Any]], float]:
    totals: dict[str, dict[str, Any]] = {
        label: {"warehouseName": label, "salesQty": 0.0, "returnQty": 0.0}
        for label in WAREHOUSE_DISPLAY_ORDER
    }
    for row in raw_rows:
        label = map_warehouse_name(row.get("rawWarehouseName"), warehouse_map)
        bucket = totals[label]
        bucket["salesQty"] += _float(row.get("salesQty"))
        bucket["returnQty"] += _float(row.get("returnQty"))

    total_return_qty = sum(_float(row["returnQty"]) for row in totals.values())
    dirty_return_qty = sum(_float(totals[label]["returnQty"]) for label in ("未映射", "无仓库记录"))
    out: list[dict[str, Any]] = []
    for label in WAREHOUSE_DISPLAY_ORDER:
        row = totals[label]
        sales_qty = _float(row["salesQty"])
        return_qty = _float(row["returnQty"])
        out.append({
            "warehouseName": label,
            "salesQty": sales_qty,
            "returnQty": return_qty,
            "returnShare": _rate(return_qty, total_return_qty),
            "returnRate": _rate(return_qty, sales_qty),
        })
    return out, (_rate(dirty_return_qty, total_return_qty) or 0.0)


def cohort_effective_age_days(metadata: dict[str, Any], since: str | None) -> int | None:
    as_of = metadata.get("asOfDate")
    covered_through = metadata.get("coveredThrough")
    if not as_of or not since or not covered_through or covered_through == "-":
        return None
    try:
        since_date = _parse_iso_date(str(since))
        covered_through_date = _parse_iso_date(str(covered_through))
        midpoint = since_date + timedelta(days=(covered_through_date - since_date).days // 2)
        return max(0, (_parse_iso_date(str(as_of)) - midpoint).days)
    except argparse.ArgumentTypeError:
        return None


def timing_training_since(metadata: dict[str, Any], timing_maturity_days: int) -> str:
    as_of = metadata.get("asOfDate")
    try:
        today = _parse_iso_date(str(as_of)) if as_of else date.today()
    except argparse.ArgumentTypeError:
        today = date.today()
    return (today - timedelta(days=timing_maturity_days + 365)).isoformat()


def rollup_style_type_rates(
    style_rows: list[dict[str, Any]],
    style_tags: dict[str, StyleTag],
) -> list[dict[str, Any]]:
    totals: dict[str, dict[str, Any]] = {
        style_type: {"styleType": style_type, "styleCount": 0, "salesQty": 0.0, "returnQty": 0.0}
        for style_type in STYLE_TYPE_ORDER
    }
    for row in style_rows:
        style_type = style_type_for(str(row.get("styleCode") or ""), style_tags)
        bucket = totals[style_type]
        bucket["styleCount"] += 1
        bucket["salesQty"] += _float(row.get("salesQty"))
        bucket["returnQty"] += _float(row.get("returnQty"))

    total_sales_qty = sum(_float(row["salesQty"]) for row in totals.values())
    out: list[dict[str, Any]] = []
    for style_type in STYLE_TYPE_ORDER:
        row = totals[style_type]
        sales_qty = _float(row["salesQty"])
        return_qty = _float(row["returnQty"])
        out.append({
            "styleType": style_type,
            "styleCount": _int(row["styleCount"]),
            "salesQty": sales_qty,
            "salesShare": _rate(sales_qty, total_sales_qty),
            "returnQty": return_qty,
            "returnRate": _rate(return_qty, sales_qty),
        })
    return out


def rollup_timing_by_style_type(
    timing_rows: list[dict[str, Any]],
    style_tags: dict[str, StyleTag],
) -> list[dict[str, Any]]:
    totals: dict[str, dict[str, Any]] = {
        style_type: {"styleType": style_type, "returnedQty": 0.0, "qty_0_30": 0.0, "qty_31_45": 0.0, "qty_45plus": 0.0}
        for style_type in STYLE_TYPE_ORDER
    }
    for row in timing_rows:
        style_type = style_type_for(str(row.get("styleCode") or ""), style_tags)
        bucket = totals[style_type]
        for key in ("returnedQty", "qty_0_30", "qty_31_45", "qty_45plus"):
            bucket[key] += _float(row.get(key))

    out: list[dict[str, Any]] = []
    for style_type in STYLE_TYPE_ORDER:
        row = totals[style_type]
        returned_qty = _float(row["returnedQty"])
        out.append({
            "styleType": style_type,
            "returnedQty": returned_qty,
            "pct_0_30": _rate(_float(row["qty_0_30"]), returned_qty),
            "pct_31_45": _rate(_float(row["qty_31_45"]), returned_qty),
            "pct_45plus": _rate(_float(row["qty_45plus"]), returned_qty),
        })
    return out


def annotate_top_styles(
    rows: list[dict[str, Any]],
    style_tags: dict[str, StyleTag],
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        style_code = str(item.get("styleCode") or "")
        item["styleType"] = style_type_for(style_code, style_tags)
        out.append(item)
    return out


def build_site_report_data(
    conn: pymysql.Connection,
    site: str,
    window: WeekWindow,
    style_tags: dict[str, StyleTag],
    top: int,
    min_qty: int,
) -> SiteReportData:
    account = site_to_account(site)
    logger.info(
        "fetching site return report site={} account={} since={} until={} maturity_days={}",
        site,
        account,
        window.since,
        window.until or "<CURDATE-maturityDays>",
        window.maturity_days,
    )
    summary, metadata = fetch_overall_summary(
        conn,
        account,
        window.since,
        window.until,
        window.maturity_days,
    )
    style_rows = fetch_site_return_rate_by_style(
        conn,
        account,
        window.since,
        window.until,
        window.maturity_days,
    )
    top_styles = fetch_site_return_rate_by_style(
        conn,
        account,
        window.since,
        window.until,
        window.maturity_days,
        top=top,
        min_qty=min_qty,
    )
    timing_style_rows = fetch_site_return_timing_by_style(
        conn,
        account,
        window.since,
        window.until,
        window.maturity_days,
    )
    order_unit_rows = fetch_site_return_rate_by_order_units(
        conn,
        account,
        window.since,
        window.until,
        window.maturity_days,
    )
    warehouse_rows, dirty_warehouse_pct = fetch_site_return_rate_by_warehouse(
        conn,
        account,
        window.since,
        window.until,
        window.maturity_days,
    )
    return SiteReportData(
        site=site,
        account=account,
        metadata=metadata,
        summary=summary,
        style_type_rows=rollup_style_type_rates(style_rows, style_tags),
        top_styles=annotate_top_styles(top_styles, style_tags),
        timing_rows=rollup_timing_by_style_type(timing_style_rows, style_tags),
        order_unit_rows=order_unit_rows,
        warehouse_rows=warehouse_rows,
        dirty_warehouse_pct=dirty_warehouse_pct,
    )


def _group_full_style_rows(
    style_rows: list[dict[str, Any]],
    style_tags: dict[str, StyleTag],
) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {style_type: [] for style_type in FULL_STYLE_TYPE_ORDER}
    for row in style_rows:
        item = dict(row)
        style_code = str(item.get("styleCode") or "")
        style_type = full_style_type_for(style_code, style_tags)
        item["styleType"] = style_type
        grouped[style_type].append(item)
    for rows in grouped.values():
        rows.sort(key=lambda row: str(row.get("styleCode") or ""))
    return grouped


def _timing_rows_by_style_type(
    timing_rows: list[dict[str, Any]],
    style_tags: dict[str, StyleTag],
) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]], dict[str, Any]]:
    by_style = {str(row.get("styleCode") or ""): row for row in timing_rows if row.get("styleCode")}
    type_source: dict[str, list[dict[str, Any]]] = {style_type: [] for style_type in FULL_STYLE_TYPE_ORDER}
    for row in timing_rows:
        style_type = full_style_type_for(str(row.get("styleCode") or ""), style_tags)
        type_source[style_type].append(row)
    by_type = {style_type: _sum_timing_rows(rows) for style_type, rows in type_source.items()}
    site_curve = _sum_timing_rows(timing_rows)
    return by_style, by_type, site_curve


def build_full_table1_rows(
    style_rows: list[dict[str, Any]],
    mature_timing_rows: list[dict[str, Any]],
    style_tags: dict[str, StyleTag],
    age_days: int | None,
    timing_maturity_days: int = DEFAULT_FULL_TIMING_MATURITY_DAYS,
) -> list[dict[str, Any]]:
    grouped = _group_full_style_rows(style_rows, style_tags)
    style_curves, type_curves, site_curve = _timing_rows_by_style_type(mature_timing_rows, style_tags)
    out: list[dict[str, Any]] = []
    for style_type in FULL_STYLE_TYPE_ORDER:
        children = grouped[style_type]
        sales_qty = sum(_float(row.get("salesQty")) for row in children)
        return_qty = sum(_float(row.get("returnQty")) for row in children)
        current_rate = _rate(return_qty, sales_qty)
        prediction = predict_restored_return_rate_beta(
            current_rate,
            age_days,
            None,
            type_curves.get(style_type),
            site_curve,
            maturity_days=timing_maturity_days,
        )
        subtotal_low_confidence = prediction.low_confidence
        if (
            prediction.curve_level == "styleType"
            and _float((type_curves.get(style_type) or {}).get("returnedQty")) >= DEFAULT_PREDICTION_MIN_RETURN_QTY
        ):
            subtotal_low_confidence = False
        out.append({
            "styleType": style_type,
            "styleCode": "小计",
            "salesQty": sales_qty,
            "returnQty": return_qty,
            "returnRate": current_rate,
            "predictedReturnRate": prediction.rate,
            "predictionProgress": prediction.progress,
            "predictionCurveLevel": prediction.curve_level,
            "lowConfidence": subtotal_low_confidence,
            "isSubtotal": True,
        })
        for row in children:
            style_code = str(row.get("styleCode") or "")
            style_sales_qty = _float(row.get("salesQty"))
            style_return_qty = _float(row.get("returnQty"))
            style_current_rate = _rate(style_return_qty, style_sales_qty)
            prediction = predict_restored_return_rate_beta(
                style_current_rate,
                age_days,
                style_curves.get(style_code),
                type_curves.get(style_type),
                site_curve,
                maturity_days=timing_maturity_days,
            )
            out.append({
                "styleType": style_type,
                "styleCode": style_code,
                "salesQty": style_sales_qty,
                "returnQty": style_return_qty,
                "returnRate": style_current_rate,
                "predictedReturnRate": prediction.rate,
                "predictionProgress": prediction.progress,
                "predictionCurveLevel": prediction.curve_level,
                "lowConfidence": prediction.low_confidence,
                "isSubtotal": False,
            })
    return out


def build_full_table2_rows(
    style_rows: list[dict[str, Any]],
    mature_timing_rows: list[dict[str, Any]],
    style_tags: dict[str, StyleTag],
) -> list[dict[str, Any]]:
    grouped = _group_full_style_rows(style_rows, style_tags)
    timing_by_style, type_curves, _site_curve = _timing_rows_by_style_type(mature_timing_rows, style_tags)
    out: list[dict[str, Any]] = []
    for style_type in FULL_STYLE_TYPE_ORDER:
        type_curve = type_curves.get(style_type) or {}
        out.append({
            "styleType": style_type,
            "styleCode": "小计",
            "pct_0_30": type_curve.get("pct_0_30"),
            "pct_31_45": type_curve.get("pct_31_45"),
            "pct_45plus": type_curve.get("pct_45plus"),
            "isSubtotal": True,
        })
        known_styles = {str(row.get("styleCode") or "") for row in grouped[style_type]}
        for row in mature_timing_rows:
            style_code = str(row.get("styleCode") or "")
            if full_style_type_for(style_code, style_tags) == style_type:
                known_styles.add(style_code)
        for style_code in sorted(style for style in known_styles if style):
            timing = timing_by_style.get(style_code) or {}
            out.append({
                "styleType": style_type,
                "styleCode": style_code,
                "pct_0_30": timing.get("pct_0_30"),
                "pct_31_45": timing.get("pct_31_45"),
                "pct_45plus": timing.get("pct_45plus"),
                "isSubtotal": False,
            })
    return out


def build_full_site_report_data(
    conn: pymysql.Connection,
    site: str,
    window: WeekWindow,
    style_tags: dict[str, StyleTag],
    warehouse_map: dict[str, str],
    timing_maturity_days: int = DEFAULT_FULL_TIMING_MATURITY_DAYS,
) -> FullSiteReportData:
    account = site_to_account(site)
    logger.info(
        "fetching full site return report site={} account={} since={} until={} maturity_days={} timing_maturity_days={}",
        site,
        account,
        window.since,
        window.until or "<CURDATE-maturityDays>",
        window.maturity_days,
        timing_maturity_days,
    )
    summary, metadata = fetch_overall_summary(
        conn,
        account,
        window.since,
        window.until,
        window.maturity_days,
    )
    style_rows = fetch_site_return_rate_by_style(
        conn,
        account,
        window.since,
        window.until,
        window.maturity_days,
    )
    timing_since = timing_training_since(metadata, timing_maturity_days)
    logger.info(
        "fetching mature timing curve site={} account={} timing_since={} timing_until=<CURDATE-timingMaturityDays> timing_maturity_days={}",
        site,
        account,
        timing_since,
        timing_maturity_days,
    )
    mature_timing_rows, timing_metadata = fetch_site_return_timing_by_style_with_metadata(
        conn,
        account,
        timing_since,
        None,
        timing_maturity_days,
    )
    order_unit_rows = fetch_site_return_rate_by_order_units(
        conn,
        account,
        window.since,
        window.until,
        window.maturity_days,
    )
    raw_warehouse_rows = fetch_site_return_rate_by_raw_warehouse(
        conn,
        account,
        window.since,
        window.until,
        window.maturity_days,
    )
    warehouse_rows, dirty_warehouse_pct = rollup_mapped_warehouse_rows(raw_warehouse_rows, warehouse_map)
    age_days = cohort_effective_age_days(metadata, window.since)
    return FullSiteReportData(
        site=site,
        account=account,
        metadata=metadata,
        summary=summary,
        cohort_age_days=age_days,
        table1_rows=build_full_table1_rows(
            style_rows,
            mature_timing_rows,
            style_tags,
            age_days,
            timing_maturity_days=timing_maturity_days,
        ),
        table2_rows=build_full_table2_rows(style_rows, mature_timing_rows, style_tags),
        order_unit_rows=order_unit_rows,
        warehouse_rows=warehouse_rows,
        dirty_warehouse_pct=dirty_warehouse_pct,
        timing_metadata=timing_metadata,
    )


def build_full_report_data(
    conn: pymysql.Connection,
    sites: list[str],
    window: WeekWindow,
    timing_maturity_days: int = DEFAULT_FULL_TIMING_MATURITY_DAYS,
) -> FullReportData:
    style_tags = load_style_tags()
    warehouse_map = load_warehouse_map()
    return FullReportData(
        since=window.since,
        until=window.until,
        maturity_days=window.maturity_days,
        timing_maturity_days=timing_maturity_days,
        sites=[
            build_full_site_report_data(
                conn,
                site,
                window,
                style_tags,
                warehouse_map,
                timing_maturity_days=timing_maturity_days,
            )
            for site in sites
        ],
    )


def build_report_data(
    conn: pymysql.Connection,
    sites: list[str],
    window: WeekWindow,
    top: int,
    min_qty: int,
) -> ReportData:
    style_tags = load_style_tags()
    return ReportData(
        since=window.since,
        until=window.until,
        maturity_days=window.maturity_days,
        sites=[
            build_site_report_data(conn, site, window, style_tags, top=top, min_qty=min_qty)
            for site in sites
        ],
    )


def _md_cell(value: Any) -> str:
    text = str(value)
    return text.replace("\n", " ").replace("|", "\\|")


def render_table(headers: list[str], rows: list[list[Any]]) -> str:
    lines = [
        "| " + " | ".join(_md_cell(h) for h in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    if rows:
        lines.extend("| " + " | ".join(_md_cell(c) for c in row) + " |" for row in rows)
    else:
        lines.append("| " + " | ".join(["-"] * len(headers)) + " |")
    return "\n".join(lines)


def covered_through_from_window_end(window_end: Any) -> str:
    if not window_end:
        return "-"
    try:
        return (_parse_iso_date(str(window_end)) - timedelta(days=1)).isoformat()
    except argparse.ArgumentTypeError:
        return str(window_end)


def inclusive_window_label(metadata: dict[str, Any], fallback_since: str | None = None) -> str:
    window_start = metadata.get("windowStart") or fallback_since or "-"
    covered_through = metadata.get("coveredThrough") or covered_through_from_window_end(metadata.get("windowEnd"))
    return f"{window_start} ~ {covered_through} (含)"


def full_order_sales_date_label(metadata: dict[str, Any], fallback_since: str | None = None) -> str:
    window_start = metadata.get("since") or metadata.get("windowStart") or fallback_since or "-"
    covered_through = metadata.get("coveredThrough") or covered_through_from_window_end(metadata.get("windowEnd"))
    return f"{window_start} ~ {covered_through}"


def full_refund_date_label(metadata: dict[str, Any]) -> str:
    return f"截至 {metadata.get('asOfDate') or '-'}"


def timing_training_window_note(metadata: dict[str, Any], timing_maturity_days: int) -> str:
    window_start = metadata.get("windowStart") or "-"
    window_end = metadata.get("windowEnd") or "-"
    covered_through = metadata.get("coveredThrough") or covered_through_from_window_end(window_end)
    return (
        f"表二/预测曲线来自成熟训练窗口 [{window_start} ~ {window_end})"
        f"（覆盖至 {covered_through} 含），maturity={timing_maturity_days}；不是展示窗口。"
    )


def _qty(value: Any) -> str:
    number = _float(value)
    if number == int(number):
        return str(int(number))
    return f"{number:.2f}"


def render_site_metadata(data: SiteReportData) -> str:
    metadata = data.metadata
    summary = data.summary
    return render_table(
        ["account", "asOfDate", "cohort窗口", "maturityDays", "salesQty", "returnQty", "current退货率"],
        [[
            data.account,
            metadata.get("asOfDate") or "-",
            inclusive_window_label(metadata),
            metadata.get("maturityDays") or "-",
            _qty(summary.get("salesQty")),
            _qty(summary.get("returnQty")),
            _pct(summary.get("returnRate")),
        ]],
    )


def render_style_type_table(rows: list[dict[str, Any]]) -> str:
    return render_table(
        ["款式类型", "款数", "销量", "销量占比", "退量", "当前还原退款率"],
        [
            [
                row["styleType"],
                row["styleCount"],
                _qty(row["salesQty"]),
                _pct(row["salesShare"]),
                _qty(row["returnQty"]),
                _pct(row["returnRate"]),
            ]
            for row in rows
        ],
    )


def render_top_style_table(rows: list[dict[str, Any]]) -> str:
    return render_table(
        ["style", "款式类型", "销量", "退量", "当前还原退款率", "SKU数"],
        [
            [
                row.get("styleCode") or "-",
                row.get("styleType") or "-",
                _qty(row.get("salesQty")),
                _qty(row.get("returnQty")),
                _pct(row.get("returnRate")),
                row.get("skuCount") or 0,
            ]
            for row in rows
        ],
    )


def render_timing_table(rows: list[dict[str, Any]]) -> str:
    return render_table(
        ["款式类型", "退量", "30天内", "31-45天", "45天以上"],
        [
            [
                row["styleType"],
                _qty(row["returnedQty"]),
                _pct(row["pct_0_30"]),
                _pct(row["pct_31_45"]),
                _pct(row["pct_45plus"]),
            ]
            for row in rows
        ],
    )


def render_order_unit_table(rows: list[dict[str, Any]]) -> str:
    return render_table(
        ["订单件数", "订单数", "销量", "退量", "当前退货率"],
        [
            [
                row.get("unitsBucket") or "-",
                row.get("orderCount") or 0,
                _qty(row.get("salesQty")),
                _qty(row.get("returnQty")),
                _pct(row.get("returnRate")),
            ]
            for row in rows
        ],
    )


def render_warehouse_table(rows: list[dict[str, Any]]) -> str:
    return render_table(
        ["发货仓", "销量", "退量", "退货占比", "当前退货率"],
        [
            [
                row.get("warehouseName") or "无仓库记录",
                _qty(row.get("salesQty")),
                _qty(row.get("returnQty")),
                _pct(row.get("returnShare")),
                _pct(row.get("returnRate")),
            ]
            for row in rows
        ],
    )


def render_site_report(data: SiteReportData, low_sample_sales_qty: int = DEFAULT_LOW_SAMPLE_SALES_QTY) -> str:
    low_sample_note = ""
    if data.site == "DE" or _float(data.summary.get("salesQty")) < low_sample_sales_qty:
        low_sample_note = (
            f"\n\n> 低样本提示：{data.site} salesQty={_qty(data.summary.get('salesQty'))}, "
            f"returnQty={_qty(data.summary.get('returnQty'))}，样本量偏小，当前退货率仅作方向参考。"
        )
    return "\n\n".join([
        f"## {data.site}",
        render_site_metadata(data) + low_sample_note,
        "**表一：还原退款率（按款式类型）**\n" + render_style_type_table(data.style_type_rows),
        "**该站高退货款 TOP10（销量≥50）**\n" + render_top_style_table(data.top_styles),
        "**表二：退货时间分布（仅退货行）**\n" + render_timing_table(data.timing_rows),
        "**表三A：订单件数分档退货率**\n" + render_order_unit_table(data.order_unit_rows),
        (
            "**表三B：发货仓退货占比与退货率**\n"
            f"dirtyWarehousePct={_pct(data.dirty_warehouse_pct)}\n"
            + (
                f"> 仓字段告警：{data.site} 仓字段缺失严重"
                f"（无仓库记录 dirtyWarehousePct={_pct(data.dirty_warehouse_pct)}），"
                "表三B 仅用于暴露仓字段缺失，不建议据此比较各仓退货率。\n"
                if data.dirty_warehouse_pct >= 0.30
                else ""
            )
            + "\n"
            + render_warehouse_table(data.warehouse_rows)
        ),
    ])


def render_markdown_report(data: ReportData) -> tuple[str, str]:
    first_site = data.sites[0] if data.sites else None
    window_label = "-"
    as_of = "-"
    if first_site:
        window_label = inclusive_window_label(first_site.metadata, data.since)
        as_of = str(first_site.metadata.get("asOfDate") or "-")
    title = f"独立站 Shopify 退货率周报 {window_label}"
    body = [
        f"# {title}",
        (
            f"- asOfDate: {as_of}\n"
            f"- maturityDays: {data.maturity_days}\n"
            "- 口径：cohort 按 pay_time；退货率 = SUM(COALESCE(return_quantity,0)) / SUM(quantity)；"
            "只有退货时间分布筛 return_quantity > 0。\n"
            "- v1：仅展示 current。"
        ),
    ]
    body.extend(render_site_report(site) for site in data.sites)
    return title, "\n\n".join(body)


def _full_style_display(row: dict[str, Any]) -> str:
    style = str(row.get("styleCode") or "-")
    if row.get("lowConfidence"):
        return f"{style} (lowConfidence)"
    return style


def _full_order_units_label(value: Any) -> str:
    bucket = str(value or "-")
    return {
        "1": "1件",
        "2": "2件",
        "3": "3件",
        "4": "4件",
        "5+": "5件以上",
    }.get(bucket, bucket)


def render_full_table1(rows: list[dict[str, Any]], order_sales_date: str, refund_date: str) -> str:
    return render_table(
        ["款式类型", "style", "订单销售日期", "退款日期", "当前还原退款率", "预测还原退款率"],
        [
            [
                row.get("styleType") or "-",
                _full_style_display(row),
                order_sales_date,
                refund_date,
                _pct(row.get("returnRate")),
                _pct(row.get("predictedReturnRate")),
            ]
            for row in rows
        ],
    )


def render_full_table2(rows: list[dict[str, Any]]) -> str:
    return render_table(
        ["款式类型", "style", "30天退货占比", "45天退货占比", "45天以上退货占比"],
        [
            [
                row.get("styleType") or "-",
                row.get("styleCode") or "-",
                _pct(row.get("pct_0_30")),
                _pct(row.get("pct_31_45")),
                _pct(row.get("pct_45plus")),
            ]
            for row in rows
        ],
    )


def render_full_order_unit_table(rows: list[dict[str, Any]], order_sales_date: str, refund_date: str) -> str:
    return render_table(
        ["订单", "订单销售日期", "退款日期", "当前还原退款率"],
        [
            [
                _full_order_units_label(row.get("unitsBucket")),
                order_sales_date,
                refund_date,
                _pct(row.get("returnRate")),
            ]
            for row in rows
        ],
    )


def render_full_warehouse_table(rows: list[dict[str, Any]]) -> str:
    return render_table(
        ["发货仓库", "退货占比", "当前还原退款率"],
        [
            [
                row.get("warehouseName") or "无仓库记录",
                _pct(row.get("returnShare")),
                _pct(row.get("returnRate")),
            ]
            for row in rows
        ],
    )


def render_full_site_report(data: FullSiteReportData) -> str:
    timing_note = timing_training_window_note(
        data.timing_metadata,
        _int(data.timing_metadata.get("maturityDays") or DEFAULT_FULL_TIMING_MATURITY_DAYS),
    )
    order_sales_date = full_order_sales_date_label(data.metadata)
    refund_date = full_refund_date_label(data.metadata)
    return "\n\n".join([
        f"## {data.site}",
        render_site_metadata(
            SiteReportData(
                site=data.site,
                account=data.account,
                metadata=data.metadata,
                summary=data.summary,
                style_type_rows=[],
                top_styles=[],
                timing_rows=[],
                order_unit_rows=[],
                warehouse_rows=[],
                dirty_warehouse_pct=data.dirty_warehouse_pct,
            )
        ),
        (
            "**表一：款式类型 × style 还原退款率**\n"
            f"> cohort有效age={data.cohort_age_days if data.cohort_age_days is not None else '-'}天；"
            "lowConfidence 表示该 style 成熟退量样本不足，预测曲线已回退到款式类型或站点。\n"
            + render_full_table1(data.table1_rows, order_sales_date, refund_date)
        ),
        "**表二：款式类型 × style 退货时间分布**\n> " + timing_note + "\n" + render_full_table2(data.table2_rows),
        "**表三A：订单件数分档退货率**\n" + render_full_order_unit_table(data.order_unit_rows, order_sales_date, refund_date),
        (
            "**表三B：发货仓退货占比与退货率**\n"
            f"dirtyWarehousePct={_pct(data.dirty_warehouse_pct)}\n\n"
            + render_full_warehouse_table(data.warehouse_rows)
        ),
    ])


def render_full_markdown_report(data: FullReportData) -> tuple[str, str]:
    first_site = data.sites[0] if data.sites else None
    window_label = "-"
    as_of = "-"
    timing_note = "表二/预测曲线来自成熟训练窗口 [- ~ -)，maturity=" + str(data.timing_maturity_days) + "；不是展示窗口。"
    if first_site:
        window_label = inclusive_window_label(first_site.metadata, data.since)
        as_of = str(first_site.metadata.get("asOfDate") or "-")
        timing_note = timing_training_window_note(first_site.timing_metadata, data.timing_maturity_days)
    title = f"独立站 Shopify 退货率明细版 {window_label}"
    body = [
        f"# {title}",
        (
            f"- asOfDate: {as_of}\n"
            f"- maturityDays: {data.maturity_days}\n"
            f"- timingMaturityDays: {data.timing_maturity_days}\n"
            f"- {timing_note}\n"
            "- 口径：cohort 按 pay_time；style=LEFT(shipping_sku,7)；"
            "分母=SUM(quantity)，分子=SUM(COALESCE(return_quantity,0))；"
            "return_time 仅用于表二退货时间分布。\n"
            "- 款式类型：site_style_tag.json；迭代款按迭代后展示；未匹配款按老款展示。\n"
            "- 预测还原退款率(beta)：当前还原退款率 / 成熟进度 progress(age)，"
            "progress floor=30.0%，预测率 cap=95.0%。"
        ),
    ]
    body.extend(render_full_site_report(site) for site in data.sites)
    return title, "\n\n".join(body)


def _xlsx_safe_sheet_name(name: str) -> str:
    return re.sub(r"[\[\]\:\*\?\/\\]", "_", name)[:31] or "Sheet"


def _xlsx_write_table(
    ws: Any,
    row_idx: int,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    percent_columns: set[int] | None = None,
) -> int:
    from openpyxl.styles import Font, PatternFill

    percent_columns = percent_columns or set()
    ws.cell(row=row_idx, column=1, value=title)
    ws.cell(row=row_idx, column=1).font = Font(bold=True, size=13)
    row_idx += 1
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
    row_idx += 1
    if not rows:
        rows = [["-" for _ in headers]]
    for values in rows:
        is_subtotal = bool(values[-1]) if len(values) == len(headers) + 1 else False
        display_values = values[:-1] if len(values) == len(headers) + 1 else values
        for col_idx, value in enumerate(display_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in percent_columns and isinstance(value, (int, float)):
                cell.number_format = "0.0%"
            if is_subtotal:
                cell.font = Font(bold=True)
        row_idx += 1
    return row_idx + 2


def export_full_report_xlsx(data: FullReportData, path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)
    for site in data.sites:
        ws = wb.create_sheet(_xlsx_safe_sheet_name(site.site))
        row_idx = 1
        order_sales_date = full_order_sales_date_label(site.metadata, data.since)
        refund_date = full_refund_date_label(site.metadata)
        ws.cell(row=row_idx, column=1, value=f"{site.site} 独立站退货率明细版")
        row_idx += 2
        meta_headers = ["account", "asOfDate", "cohort窗口", "maturityDays", "cohort有效age", "salesQty", "returnQty", "current退货率"]
        meta_rows = [[
            site.account,
            site.metadata.get("asOfDate") or "-",
            inclusive_window_label(site.metadata),
            site.metadata.get("maturityDays") or "-",
            site.cohort_age_days if site.cohort_age_days is not None else "-",
            _float(site.summary.get("salesQty")),
            _float(site.summary.get("returnQty")),
            site.summary.get("returnRate"),
        ]]
        row_idx = _xlsx_write_table(ws, row_idx, "概览", meta_headers, meta_rows, percent_columns={8})
        table1_rows = [
            [
                row.get("styleType") or "-",
                _full_style_display(row),
                order_sales_date,
                refund_date,
                row.get("returnRate"),
                row.get("predictedReturnRate"),
                row.get("isSubtotal"),
            ]
            for row in site.table1_rows
        ]
        row_idx = _xlsx_write_table(
            ws,
            row_idx,
            "表一：款式类型 × style 还原退款率",
            ["款式类型", "style", "订单销售日期", "退款日期", "当前还原退款率", "预测还原退款率"],
            table1_rows,
            percent_columns={5, 6},
        )
        table2_rows = [
            [
                row.get("styleType") or "-",
                row.get("styleCode") or "-",
                row.get("pct_0_30"),
                row.get("pct_31_45"),
                row.get("pct_45plus"),
                row.get("isSubtotal"),
            ]
            for row in site.table2_rows
        ]
        row_idx = _xlsx_write_table(
            ws,
            row_idx,
            "表二：款式类型 × style 退货时间分布",
            ["款式类型", "style", "30天退货占比", "45天退货占比", "45天以上退货占比"],
            table2_rows,
            percent_columns={3, 4, 5},
        )
        order_rows = [
            [
                _full_order_units_label(row.get("unitsBucket")),
                order_sales_date,
                refund_date,
                row.get("returnRate"),
            ]
            for row in site.order_unit_rows
        ]
        row_idx = _xlsx_write_table(
            ws,
            row_idx,
            "表三A：订单件数分档退货率",
            ["订单", "订单销售日期", "退款日期", "当前还原退款率"],
            order_rows,
            percent_columns={4},
        )
        warehouse_rows = [
            [
                row.get("warehouseName") or "-",
                row.get("returnShare"),
                row.get("returnRate"),
            ]
            for row in site.warehouse_rows
        ]
        ws.cell(row=row_idx, column=1, value=f"dirtyWarehousePct={_pct(site.dirty_warehouse_pct)}")
        row_idx += 1
        _xlsx_write_table(
            ws,
            row_idx,
            "表三B：发货仓退货占比与退货率",
            ["发货仓库", "退货占比", "当前还原退款率"],
            warehouse_rows,
            percent_columns={2, 3},
        )
        for column_cells in ws.columns:
            max_len = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                value = cell.value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[column].width = min(max(max_len + 2, 10), 36)
    wb.save(path)


def write_full_report_outputs(data: FullReportData, markdown: str, output_dir: Path = OUTPUT_DIR) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    first_site = data.sites[0] if data.sites else None
    as_of = str(first_site.metadata.get("asOfDate") if first_site else date.today().isoformat())
    covered = str(first_site.metadata.get("coveredThrough") if first_site else data.since)
    stem = f"site_return_full_{as_of}_{data.since}_to_{covered}"
    md_path = output_dir / f"{stem}.md"
    xlsx_path = output_dir / f"{stem}.xlsx"
    md_path.write_text(markdown, encoding="utf-8")
    export_full_report_xlsx(data, xlsx_path)
    return md_path, xlsx_path


def _iter_conversation_records(value: Any) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    if isinstance(value, list):
        for item in value:
            if isinstance(item, dict):
                records.append(item)
    elif isinstance(value, dict):
        if any(key in value for key in ("openConversationId", "open_conversation_id", "conversationId", "id")):
            records.append(value)
        for key in ("conversations", "items", "groups", "data", "result"):
            records.extend(_iter_conversation_records(value.get(key)))
        for key, item in value.items():
            if isinstance(item, str):
                records.append({"name": key, "openConversationId": item})
            elif isinstance(item, dict):
                merged = {"name": key, **item} if "name" not in item else item
                records.extend(_iter_conversation_records(merged))
    return records


def lookup_open_conversation_id(path: Path, group_name: str) -> str:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise LookupError(f"conversation file not found: {path}") from exc
    except json.JSONDecodeError as exc:
        raise LookupError(f"conversation file is invalid JSON: {path}: {exc}") from exc

    for record in _iter_conversation_records(payload):
        names = [
            record.get("name"),
            record.get("title"),
            record.get("groupName"),
            record.get("conversationName"),
            record.get("chatName"),
        ]
        if group_name not in {str(name) for name in names if name}:
            continue
        conversation_id = (
            record.get("openConversationId")
            or record.get("open_conversation_id")
            or record.get("conversationId")
            or record.get("id")
        )
        if conversation_id:
            return str(conversation_id)

    raise LookupError(f"group {group_name!r} not found in {path}")


def resolve_dingtalk_target(
    channel: str,
    group_name: str | None,
    channels_file: Path,
    conversations_file: Path,
) -> DingTalkTarget:
    channels = _load_json(channels_file)
    entry = channels.get(channel, {}) if isinstance(channels, dict) else {}
    if not isinstance(entry, dict):
        entry = {}

    secrets = _load_tool_secrets()
    plist_env: dict[str, Any] = {}
    bot_plist = Path(os.path.expanduser("~/Library/LaunchAgents/com.everpretty.dingtalk-bot.plist"))
    try:
        with bot_plist.open("rb") as f:
            plist_env = plistlib.load(f).get("EnvironmentVariables", {})
    except FileNotFoundError:
        plist_env = {}
    bot_dotenv = _read_dotenv(Path(os.path.expanduser("~/PycharmProjects/paperclip-dingtalk-bot/.env")))

    app_key = (
        os.environ.get("DINGTALK_APP_KEY")
        or entry.get("app_key")
        or entry.get("appKey")
        or plist_env.get("DINGTALK_APP_KEY")
        or bot_dotenv.get("DINGTALK_APP_KEY")
        or _find_normalized(secrets, {"dingtalk_app_key"})
        or _find_in_named_section(secrets, "dingtalk", {"app_key", "appkey", "key"})
    )
    app_secret = (
        os.environ.get("DINGTALK_APP_SECRET")
        or entry.get("app_secret")
        or entry.get("appSecret")
        or plist_env.get("DINGTALK_APP_SECRET")
        or bot_dotenv.get("DINGTALK_APP_SECRET")
        or _find_normalized(secrets, {"dingtalk_app_secret"})
        or _find_in_named_section(secrets, "dingtalk", {"app_secret", "appsecret", "secret"})
    )
    if not app_key or not app_secret:
        raise RuntimeError(
            "missing DingTalk credentials: set DINGTALK_APP_KEY/DINGTALK_APP_SECRET "
            f"or configure channel {channel!r} in {channels_file}"
        )

    conversation_id = (
        entry.get("conv_id")
        or entry.get("conversation_id")
        or entry.get("openConversationId")
        or entry.get("open_conversation_id")
        or entry.get("id")
    )
    if not conversation_id:
        if not group_name:
            raise RuntimeError(
                f"channel {channel!r} in {channels_file} has no conv_id/openConversationId; pass --group-name"
            )
        conversation_id = lookup_open_conversation_id(conversations_file, group_name)
    robot_code = entry.get("robot_code") or entry.get("robotCode") or str(app_key)
    return DingTalkTarget(
        app_key=str(app_key),
        app_secret=str(app_secret),
        open_conversation_id=str(conversation_id),
        robot_code=str(robot_code) if robot_code else None,
    )


class DingTalkClient:
    def __init__(
        self,
        app_key: str,
        app_secret: str,
        timeout: float = 15.0,
        client: httpx.Client | None = None,
    ) -> None:
        self.app_key = app_key
        self.app_secret = app_secret
        self.timeout = timeout
        self._client = client

    def _request_client(self) -> httpx.Client:
        if self._client is not None:
            return self._client
        return httpx.Client(timeout=self.timeout)

    def _get_access_token(self, client: httpx.Client) -> str:
        response = client.post(
            TOKEN_URL,
            json={"appKey": self.app_key, "appSecret": self.app_secret},
        )
        if response.status_code >= 300:
            raise RuntimeError(f"DingTalk token failed: {response.status_code} {response.text}")
        data = response.json()
        token = data.get("accessToken") or data.get("access_token")
        if not token:
            raise RuntimeError(f"DingTalk token response missing accessToken: {data}")
        return str(token)

    def send_markdown(
        self,
        open_conversation_id: str,
        title: str,
        text: str,
        robot_code: str | None = None,
    ) -> dict[str, Any]:
        close_client = self._client is None
        client = self._request_client()
        try:
            access_token = self._get_access_token(client)
            body = {
                "openConversationId": open_conversation_id,
                "robotCode": robot_code or self.app_key,
                "msgKey": "sampleMarkdown",
                "msgParam": json.dumps(
                    {"title": title, "text": text},
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
            }
            response = client.post(
                SEND_URL,
                headers={"x-acs-dingtalk-access-token": access_token},
                json=body,
            )
            if response.status_code >= 300:
                raise RuntimeError(f"DingTalk send failed: {response.status_code} {response.text}")
            data = response.json() if response.content else {}
            errcode = data.get("errcode")
            if errcode not in (None, 0, "0"):
                raise RuntimeError(f"DingTalk send failed: {data}")
            return data
        finally:
            if close_client:
                client.close()


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--since", help="inclusive cohort start date, YYYY-MM-DD")
    parser.add_argument("--until", help="exclusive cohort end date, YYYY-MM-DD; default DB CURDATE()-maturityDays")
    parser.add_argument("--maturity-days", type=int, default=DEFAULT_MATURITY_DAYS)
    parser.add_argument(
        "--sites",
        default=",".join(DEFAULT_SITES),
        help="comma-separated site list; default UK,US,FR,DE",
    )
    parser.add_argument("--top", type=int, default=DEFAULT_TOP_STYLES, help="high-return style limit")
    parser.add_argument("--min-qty", type=int, default=DEFAULT_MIN_TOP_STYLE_QTY, help="min salesQty for top styles")
    parser.add_argument("--full", action="store_true", help="write detailed markdown+xlsx report to scripts/output")
    parser.add_argument("--dry-run", action="store_true", help="print markdown only; this is the default")
    parser.add_argument("--send", action="store_true", help="push to DingTalk instead of only printing markdown")
    parser.add_argument("--channel", default=DEFAULT_CHANNEL, help="DingTalk channel in dingtalk-channels.json")
    parser.add_argument("--group-name", help="DingTalk group name if channel lacks conv_id")
    parser.add_argument(
        "--channels-file",
        default="~/.paperclip/dingtalk-channels.json",
        help="DingTalk channel registry JSON",
    )
    parser.add_argument(
        "--conversations-file",
        default="~/.paperclip/dingtalk_conversations.json",
        help="fallback DingTalk group openConversationId registry JSON",
    )
    return parser.parse_args(argv)


def _parse_sites(value: str) -> list[str]:
    sites = [part.strip().upper() for part in value.split(",") if part.strip()]
    if not sites:
        raise ValueError("--sites must include at least one site")
    for site in sites:
        site_to_account(site)
    return sites


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        sites = _parse_sites(args.sites)
        window = resolve_window(args.since, args.until, args.maturity_days)
        if args.maturity_days <= 0:
            raise ValueError("--maturity-days must be positive")
        if args.top <= 0:
            raise ValueError("--top must be positive")
        if args.min_qty < 0:
            raise ValueError("--min-qty must be non-negative")
    except (ValueError, argparse.ArgumentTypeError) as exc:
        logger.error("{}", exc)
        return 2

    try:
        conn = _connect()
    except Exception as exc:
        logger.error("DWS connect failed: {}", exc)
        return 2

    if args.full:
        try:
            full_data = build_full_report_data(
                conn,
                sites=sites,
                window=window,
            )
            _title, markdown = render_full_markdown_report(full_data)
            md_path, xlsx_path = write_full_report_outputs(full_data, markdown)
        except Exception as exc:
            logger.exception("full site return report build failed: {}", exc)
            return 1
        finally:
            conn.close()
        print(f"明细版 Markdown: {md_path}")
        print(f"明细版 XLSX: {xlsx_path}")
        if args.send:
            logger.warning("--full writes files only; DingTalk push is skipped")
        return 0

    try:
        data = build_report_data(
            conn,
            sites=sites,
            window=window,
            top=args.top,
            min_qty=args.min_qty,
        )
    except Exception as exc:
        logger.exception("weekly site return report build failed: {}", exc)
        return 1
    finally:
        conn.close()

    title, markdown = render_markdown_report(data)
    if not args.send:
        print(markdown)
        return 0

    try:
        target = resolve_dingtalk_target(
            channel=args.channel,
            group_name=args.group_name,
            channels_file=Path(os.path.expanduser(args.channels_file)),
            conversations_file=Path(os.path.expanduser(args.conversations_file)),
        )
    except Exception as exc:
        logger.error("DingTalk target resolution failed: {}", exc)
        return 1

    try:
        result = DingTalkClient(target.app_key, target.app_secret).send_markdown(
            target.open_conversation_id,
            title,
            markdown,
            robot_code=target.robot_code,
        )
    except Exception as exc:
        logger.error("DingTalk push failed: {}", exc)
        return 1

    logger.info("DingTalk weekly site return report sent: {}", result)
    return 0


if __name__ == "__main__":
    sys.exit(main())
