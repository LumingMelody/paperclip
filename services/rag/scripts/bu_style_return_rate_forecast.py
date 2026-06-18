#!/usr/bin/env python3
"""BU × 款号 还原退货率预测（成熟月观测口径）。

为每个 BU 列出其全部款号 + 预测退货率，输出多 Sheet Excel（每 BU 一 Sheet）。
是 bu_return_rate_forecast.py（只到 BU 级）的款号级下钻版。

口径（与 dws.returnRateByStyle / bu_return_rate_forecast.py 同源，用户确认「成熟月观测」）：
- 每个款号预测退货率 = 最近 N 个成熟销售月池化的 Σ退货件 / Σ销量件（天然按件数加权）。
- Amazon：dws_od_amazon_refund_rate_d，GROUP BY sku_left7，yearmouth ∈ 成熟月。
  （已实证 quantity ≈ 真实销量：AmazonEPUS Feb-Apr Σquantity 152,821 ≈ 订单表 152,489；
   单款 EE02960 退货表口径 54.3% == 订单表真实口径 54.4%，分母正确，礼服品类退货率本就高。）
- 独立站：dm_od_shopify_resreturn_d，GROUP BY LEFT(shipping_sku,7)，pay_time 落在成熟月。
  （退货率分子分母都取自 resreturn，自洽口径；该表约覆盖 ~68% 销量，但比率自洽。）
- BU = ods_sp_me_platform_account_m.financePlatform；BU 内跨店同款合并（件数相加）。
- 成熟月 = 月末 + maturityDays <= today 的最近 N 个完整月（默认 N=3、maturity=45 → Feb/Mar/Apr）。
  「成熟月池化销量」是这 N 个月之和，非单月——勿与单月财务销量直接对比。
- 第二张表「还原退货预测」：成熟月之后、当前月之前的完整月（未成熟，如 5 月）实际销量 ×
  成熟退货率 → 预测最终退货件/退货率（成熟率优先用该款，样本不足回退 BU 池化率）。
- shein（退货自 2025-09 断更）/ walmart / eccang(TikTok) / b2c 无款级退货源 → 标记「不可算」。

只读分析脚本，不写库。凭据从 tool-secrets.json 读。

用法：
    uv run --project services/rag python services/rag/scripts/bu_style_return_rate_forecast.py
    # 可选：--maturity-days 45 --mature-months 3 --min-qty 50 --out <path.xlsx>
"""
from __future__ import annotations

import argparse
import json
import sys
from calendar import monthrange
from datetime import date
from pathlib import Path

import pymysql

# 复用站点报告的 xlsx 写表器 + 款式标签（同口径同分类）
sys.path.insert(0, str(Path(__file__).parent))
from weekly_site_return_report import (  # noqa: E402
    FULL_STYLE_TYPE_ORDER,
    OUTPUT_DIR,
    _xlsx_write_table,
    full_style_type_for,
    load_style_tags,
)

SECRETS = Path("/Users/melodylu/.paperclip/tool-secrets.json")
DEFAULT_MATURITY_DAYS = 45
DEFAULT_MATURE_MONTHS = 3
DEFAULT_MIN_QTY = 50  # 低于此件数的款标「样本不足」（不剔除，仅提示）

# 无款级退货数据源的平台 → 不可算原因
UNCOMPUTABLE_REASON = {
    "shein": "上游 shein 退货自 2025-09 断更，不可算",
    "walmart": "无款级退货数据源，不可算",
    "eccang": "无款级退货数据源（TikTok），不可算",
    "b2c": "无款级退货数据源，不可算",
}


def _dws_conn() -> pymysql.Connection:
    s = json.loads(SECRETS.read_text(encoding="utf-8"))
    dws = list(s["companies"].values())[0]["dws"]
    return pymysql.connect(
        host=dws["host"], port=int(dws.get("port") or 3306), user=dws["user"],
        password=dws["password"], database=dws["database"], charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor, connect_timeout=8, read_timeout=120,
    )


def recent_mature_months(today: date, maturity_days: int, n: int) -> tuple[list[str], date, date]:
    """返回最近 n 个成熟月的 'YYYY-MM' 列表（旧→新），以及覆盖窗口 [start, end_exclusive)。

    月 M 成熟 ⇔ 月末日 + maturity_days <= today。
    """
    months: list[str] = []
    y, m = today.year, today.month
    # 从当前月往前找，跳过未成熟月
    while len(months) < n and (y, m) > (2000, 1):
        last_day = date(y, m, monthrange(y, m)[1])
        if (today - last_day).days >= maturity_days:
            months.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    months.reverse()
    if not months:
        raise SystemExit("no mature months found — check today/maturity_days")
    sy, sm = int(months[0][:4]), int(months[0][5:7])
    start = date(sy, sm, 1)
    ey, em = int(months[-1][:4]), int(months[-1][5:7])
    end_excl = date(ey + (em == 12), (em % 12) + 1, 1)
    return months, start, end_excl


def immature_complete_months(today: date, mature_months: list[str]) -> tuple[list[str], date, date]:
    """返回「成熟月之后、当前月之前」的完整月（旧→新）+ 覆盖窗口 [start, end_excl)。

    这些月销量已知但退货还在回流（未成熟），用于「还原退货预测」：实际销量 × 成熟率。
    当前月本身不含（销量也未跑完，预测会失真）。
    """
    if not mature_months:
        return [], today, today
    y, m = int(mature_months[-1][:4]), int(mature_months[-1][5:7])
    months: list[str] = []
    while True:
        m += 1
        if m == 13:
            y, m = y + 1, 1
        if (y, m) >= (today.year, today.month):
            break
        months.append(f"{y:04d}-{m:02d}")
    if not months:
        return [], today, today
    sy, sm = int(months[0][:4]), int(months[0][5:7])
    start = date(sy, sm, 1)
    ey, em = int(months[-1][:4]), int(months[-1][5:7])
    end_excl = date(ey + (em == 12), (em % 12) + 1, 1)
    return months, start, end_excl


def load_bu_map(conn: pymysql.Connection) -> list[dict]:
    """每个店一条：{userAccount, accountId, bu, platform}。"""
    sql = (
        "SELECT userAccount, accountId, financePlatform AS bu, LOWER(platform) AS platform "
        "FROM ods_sp_me_platform_account_m "
        "WHERE financePlatform IS NOT NULL AND financePlatform <> ''"
    )
    out: list[dict] = []
    with conn.cursor() as c:
        c.execute(sql)
        for r in c.fetchall():
            out.append({
                "userAccount": r["userAccount"],
                "accountId": r["accountId"],
                "bu": str(r["bu"]).strip(),
                "platform": (r["platform"] or "").strip(),
            })
    return out


def amazon_style_rows(conn: pymysql.Connection, account_id: int, months: list[str]) -> dict[str, dict]:
    placeholders = ",".join(["%s"] * len(months))
    sql = (
        "SELECT sku_left7 AS style, "
        "CAST(COALESCE(SUM(quantity),0) AS SIGNED) qty, "
        "CAST(COALESCE(SUM(rf_quantity),0) AS SIGNED) rf "
        "FROM dws_od_amazon_refund_rate_d "
        f"WHERE accountId=%s AND yearmouth IN ({placeholders}) "
        "AND sku_left7 IS NOT NULL AND sku_left7<>'' "
        "GROUP BY sku_left7"
    )
    out: dict[str, dict] = {}
    with conn.cursor() as c:
        c.execute(sql, (account_id, *months))
        for r in c.fetchall():
            out[r["style"]] = {"qty": int(r["qty"]), "rf": int(r["rf"])}
    return out


def shopify_style_rows(conn: pymysql.Connection, account: str, start: date, end_excl: date) -> dict[str, dict]:
    sql = (
        "SELECT LEFT(shipping_sku,7) AS style, "
        "CAST(COALESCE(SUM(quantity),0) AS SIGNED) qty, "
        "CAST(COALESCE(SUM(COALESCE(return_quantity,0)),0) AS SIGNED) rf "
        "FROM dm_od_shopify_resreturn_d "
        "WHERE account=%s AND pay_time>=%s AND pay_time<%s "
        "AND shipping_sku IS NOT NULL AND shipping_sku<>'' "
        "GROUP BY LEFT(shipping_sku,7)"
    )
    out: dict[str, dict] = {}
    with conn.cursor() as c:
        c.execute(sql, (account, start.isoformat(), end_excl.isoformat()))
        for r in c.fetchall():
            style = (r["style"] or "").strip()
            if not style:
                continue
            out[style] = {"qty": int(r["qty"]), "rf": int(r["rf"])}
    return out


def merge_into(target: dict[str, dict], src: dict[str, dict]) -> None:
    for style, v in src.items():
        t = target.setdefault(style, {"qty": 0, "rf": 0})
        t["qty"] += v["qty"]
        t["rf"] += v["rf"]


def build_bu_table_rows(style_map: dict[str, dict], style_tags, min_qty: int) -> list[list]:
    """按款式类型分组，每组先小计行后款号行（款号按销量降序）。

    返回 _xlsx_write_table 期望的行：每行末尾附 isSubtotal 标记。
    列：款式类型 / 款号 / 成熟月销量(件) / 退货量(件) / 预测退货率
    """
    by_type: dict[str, list[tuple[str, dict]]] = {}
    for style, v in style_map.items():
        st = full_style_type_for(style, style_tags)
        by_type.setdefault(st, []).append((style, v))

    type_order = list(FULL_STYLE_TYPE_ORDER) + [t for t in by_type if t not in FULL_STYLE_TYPE_ORDER]
    rows: list[list] = []
    for st in type_order:
        items = by_type.get(st)
        if not items:
            continue
        items.sort(key=lambda kv: (-kv[1]["qty"], kv[0]))
        sub_qty = sum(v["qty"] for _, v in items)
        sub_rf = sum(v["rf"] for _, v in items)
        sub_rate = (sub_rf / sub_qty) if sub_qty > 0 else None
        rows.append([st, "小计", sub_qty, sub_rf, sub_rate, True])
        for style, v in items:
            rate = (v["rf"] / v["qty"]) if v["qty"] > 0 else None
            label = style if v["qty"] >= min_qty else f"{style} (样本不足)"
            rows.append([st, label, v["qty"], v["rf"], rate, False])
    return rows


def build_restoration_table_rows(
    mature_map: dict[str, dict], imm_map: dict[str, dict], style_tags,
    bu_rate: float | None, min_qty: int,
) -> list[list]:
    """未成熟月「还原退货预测」：实际销量 × 成熟退货率（回填仍在回流的退货）。

    成熟率优先用该款 Σ退货/Σ销量（成熟池样本 >= min_qty），否则回退 BU 池化率。
    列：款式类型 / 款号 / 销量(件) / 已观测退货(件) / 已观测退货率 / 还原退货率 / 预测最终退货(件)
    """
    by_type: dict[str, list[tuple[str, dict]]] = {}
    for style, v in imm_map.items():
        st = full_style_type_for(style, style_tags)
        by_type.setdefault(st, []).append((style, v))

    type_order = list(FULL_STYLE_TYPE_ORDER) + [t for t in by_type if t not in FULL_STYLE_TYPE_ORDER]
    rows: list[list] = []
    for st in type_order:
        items = by_type.get(st)
        if not items:
            continue
        items.sort(key=lambda kv: (-kv[1]["qty"], kv[0]))
        computed: list[tuple[str, int, int, float | None, int]] = []
        for style, v in items:
            qty, obs_rf = v["qty"], v["rf"]
            mat = mature_map.get(style)
            if mat and mat["qty"] >= min_qty:
                rate = mat["rf"] / mat["qty"]
            else:
                rate = bu_rate  # 成熟样本不足 → 回退 BU 池化率
            pred_rf = round(qty * rate) if rate is not None else 0
            computed.append((style, qty, obs_rf, rate, pred_rf))
        sub_qty = sum(c[1] for c in computed)
        sub_obs = sum(c[2] for c in computed)
        sub_pred = sum(c[4] for c in computed)
        sub_obs_rate = (sub_obs / sub_qty) if sub_qty > 0 else None
        sub_pred_rate = (sub_pred / sub_qty) if sub_qty > 0 else None
        rows.append([st, "小计", sub_qty, sub_obs, sub_obs_rate, sub_pred_rate, sub_pred, True])
        for style, qty, obs_rf, rate, pred_rf in computed:
            label = style if qty >= min_qty else f"{style} (样本不足)"
            obs_rate = (obs_rf / qty) if qty > 0 else None
            rows.append([st, label, qty, obs_rf, obs_rate, rate, pred_rf, False])
    return rows


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="BU × 款号 还原退货率预测（成熟月观测口径）")
    parser.add_argument("--maturity-days", type=int, default=DEFAULT_MATURITY_DAYS)
    parser.add_argument("--mature-months", type=int, default=DEFAULT_MATURE_MONTHS,
                        help="池化最近 N 个成熟月（默认 3）")
    parser.add_argument("--min-qty", type=int, default=DEFAULT_MIN_QTY,
                        help="低于此成熟月件数的款标「样本不足」")
    parser.add_argument("--out", type=Path, default=None, help="xlsx 输出路径")
    args = parser.parse_args(argv)

    today = date.today()
    months, start, end_excl = recent_mature_months(today, args.maturity_days, args.mature_months)
    window_label = f"{months[0]} ~ {months[-1]}（{start.isoformat()} ≤ 销售月 < {end_excl.isoformat()}）"
    imm_months, imm_start, imm_end = immature_complete_months(today, months)
    pool_label = f"{months[0]}~{months[-1]} 池化"
    if len(imm_months) == 1:
        imm_label = imm_months[0]
    elif imm_months:
        imm_label = f"{imm_months[0]}~{imm_months[-1]}"
    else:
        imm_label = ""

    style_tags = load_style_tags()
    conn = _dws_conn()
    bu_map = load_bu_map(conn)

    # BU → {style_map, computable_stores, excluded_stores(原因)}
    bus: dict[str, dict] = {}
    for store in bu_map:
        bu = bus.setdefault(store["bu"], {"styles": {}, "imm_styles": {}, "computable": [], "excluded": {}})
        plat = store["platform"]
        acc = store["userAccount"]
        try:
            if "amazon" in plat:
                if store["accountId"] is None:
                    bu["excluded"][acc] = "缺 accountId"
                    continue
                aid = int(store["accountId"])
                rows = amazon_style_rows(conn, aid, months)
                imm_rows = amazon_style_rows(conn, aid, imm_months) if imm_months else {}
                plat_label = "amazon"
            elif plat == "shopify":
                rows = shopify_style_rows(conn, acc, start, end_excl)
                imm_rows = shopify_style_rows(conn, acc, imm_start, imm_end) if imm_months else {}
                plat_label = "shopify"
            else:
                bu["excluded"][acc] = UNCOMPUTABLE_REASON.get(plat, f"平台 {plat} 无款级退货源")
                continue
        except Exception as exc:  # noqa: BLE001
            bu["excluded"][acc] = f"查询失败: {exc}"
            continue
        total_qty = sum(v["qty"] for v in rows.values())
        if total_qty <= 0:
            bu["excluded"][acc] = f"{plat_label} 成熟月无销量/退货数据"
            continue
        merge_into(bu["styles"], rows)
        merge_into(bu["imm_styles"], imm_rows)
        bu["computable"].append((acc, plat_label, total_qty))
    conn.close()

    # ---- 输出 ----
    from openpyxl import Workbook

    out_path = args.out or (OUTPUT_DIR / f"bu_style_return_{today.isoformat()}.xlsx")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    print(f"# BU × 款号 还原退货率预测（成熟月观测口径）")
    print(f"- asOf: {today.isoformat()}  | 成熟月: {window_label}  | maturityDays={args.maturity_days}\n")
    print("| BU | 可算店 | 款号数 | 成熟月池化销量(件) | 退货(件) | BU池化退货率 | 不可算店 |")
    print("|----|------|------:|------:|------:|------:|------|")

    for bu in sorted(bus):
        info = bus[bu]
        style_map = info["styles"]
        bu_qty = sum(v["qty"] for v in style_map.values())
        bu_rf = sum(v["rf"] for v in style_map.values())
        bu_rate = (bu_rf / bu_qty) if bu_qty > 0 else None
        excluded_note = "；".join(f"{a}:{r}" for a, r in info["excluded"].items())

        rate_disp = f"{bu_rate*100:.1f}%" if bu_rate is not None else "—"
        comp_disp = ",".join(a for a, _, _ in info["computable"]) or "（无）"
        print(f"| {bu} | {comp_disp} | {len(style_map)} | {bu_qty:,} | {bu_rf:,} | {rate_disp} | {excluded_note or '—'} |")

        ws = wb.create_sheet(bu[:31])
        row_idx = 1
        ws.cell(row=row_idx, column=1, value=f"{bu} 款号级还原退货率预测（成熟月观测口径）")
        row_idx += 1
        if imm_months:
            note = (f"注：「成熟月池化销量」=最近 {args.mature_months} 个成熟月（{pool_label}）销量加总，非单月；"
                    f"{imm_label} 起未成熟（退货未回齐、观测率偏低）不计入观测——其还原预测见下方第二张表。")
        else:
            note = f"注：「成熟月池化销量」=最近 {args.mature_months} 个成熟月（{pool_label}）销量加总，非单月。"
        ws.cell(row=row_idx, column=1, value=note)
        row_idx += 2
        if not style_map:
            ws.cell(row=row_idx, column=1, value="本 BU 无可算款级退货数据。")
            row_idx += 1
            ws.cell(row=row_idx, column=1, value=f"不可算明细：{excluded_note or '—'}")
            continue

        meta_headers = ["BU", "口径", "成熟月窗口", "asOf", "成熟月池化销量(件)", "退货量(件)",
                        "BU成熟池化退货率", "款号数", "可算店", "不可算店"]
        meta_rows = [[
            bu, "成熟月观测 Σ退货件/Σ销量件", window_label, today.isoformat(),
            bu_qty, bu_rf, bu_rate, len(style_map),
            ",".join(a for a, _, _ in info["computable"]) or "—",
            excluded_note or "—",
        ]]
        row_idx = _xlsx_write_table(ws, row_idx, "概览", meta_headers, meta_rows, percent_columns={7})

        table_rows = build_bu_table_rows(style_map, style_tags, args.min_qty)
        row_idx = _xlsx_write_table(
            ws, row_idx,
            f"成熟月观测：款式类型 × 款号 退货率（{pool_label}，按销量降序，每组先小计）",
            ["款式类型", "款号", "成熟月池化销量(件)", "退货量(件)", "退货率"],
            table_rows,
            percent_columns={5},
        )

        imm_map = info["imm_styles"]
        if imm_months and imm_map:
            rest_rows = build_restoration_table_rows(style_map, imm_map, style_tags, bu_rate, args.min_qty)
            _xlsx_write_table(
                ws, row_idx,
                f"还原退货预测：{imm_label} 实际销量 × 成熟退货率（未成熟月退货未回齐，按销量降序，每组先小计）",
                ["款式类型", "款号", f"{imm_label}销量(件)", "已观测退货(件)",
                 "已观测退货率", "还原退货率", "预测最终退货(件)"],
                rest_rows,
                percent_columns={5, 6},
            )
        for column_cells in ws.columns:
            max_len = 0
            col = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col].width = min(max(max_len + 2, 10), 40)

    if not wb.sheetnames:
        wb.create_sheet("空")
    wb.save(out_path)
    print(f"\n→ Excel: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
